from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

from selenium import webdriver
import openpyxl
import csv
import smtplib, ssl
from mailer import Mailer
import os
import smtplib
import pandas as pd
import imghdr
from email.message import EmailMessage
from unidecode import unidecode
from datetime import datetime

driverPath = "drivers/geckodriver.exe"  # Insert your drivers path C://Windows
mime_types = "application/pdf,application/vnd.adobe.xfdf,application/vnd.fdf,application/vnd.adobe.xdp+xml"

fp = webdriver.FirefoxProfile()
fp.set_preference("browser.download.folderList", 2)
fp.set_preference("browser.download.manager.showWhenStarting", False)
fp.set_preference("plugin.disable_full_page_plugin_for_types", "application/pdf")
fp.set_preference("browser.helperApps.neverAsk.saveToDisk", mime_types)
fp.set_preference("plugin.disable_full_page_plugin_for_types", mime_types)
fp.set_preference("pdfjs.disabled", True)
fp.set_preference("javascript.enabled", False)

options = webdriver.FirefoxOptions()
# options.headless = True
options.add_argument("--window-size=1920,1080")
options.add_argument('--ignore-certificate-errors')
options.add_argument("--disable-popup-blocking")
options.add_argument("--disable-notifications")
options.add_argument('--allow-running-insecure-content')
options.add_argument("--disable-extensions")
options.add_argument("--proxy-server='direct://'")
options.add_argument("--proxy-bypass-list=*")
options.add_argument("--start-maximized")
options.add_argument('--disable-gpu')
options.add_argument('--disable-dev-shm-usage')
options.add_argument('--no-sandbox')


def filter_nonprintable(text):
    return ''.join([i if ord(i) < 128 else ' ' for i in text])


def getElements(row, i, url):
    ticker = ""
    fund = ""
    form = ""
    purpose = ""
    companyName = ""
    link = row.get_attribute("href")
    timeStamp = datetime.strptime(str(datetime.now())[:19], '%Y-%m-%d %H:%M:%S').strftime('%d-%m-%Y %I:%M:%S %p')
    emailBody = ""
    longText = ""

    try:
        form = row.find_element_by_xpath(
            '//*[@id="main-col"]/div[2]/div/div[2]/div/div/div[3]/div/a['+str(i)+"]/div[1]/div").text
        if "13D" in form: form = "13D"
        if "13G" in form: form = "13G"
        if "13F" in form: form = "13F"
    except:
        pass
    try:
        ticker = row.find_element_by_class_name("company-ticker").text
    except:
        pass
    try:
        companyName = row.find_element_by_class_name("company-name").text
    except:
        pass
    try:
        fund = row.find_element_by_class_name("company-name").text
    except:
        pass
    try:
        emailBody = row.find_element_by_xpath(
            '//*[@id="main-col"]/div[2]/div/div[2]/div/div/div[3]/div/a['+str(i)+']/div[2]/div[2]/div').text
    except:
        pass
    try:
        if url == 1:
            purpose = "Materially Definitive Agreement"
        elif url == 2:
            purpose = "Stock Split"
        elif url == 3 or url == 4:
            if ticker == "":
                purpose = "Ticker Missing"
    except:
        pass

    try:
        longText = row.find_element_by_xpath(
            '//*[@id="main-col"]/div[2]/div/div[2]/div/div/div[3]/div/a['+str(i)+']/div[2]/div[3]/div').text
    except:
        pass
    link = row.get_attribute("href")

    if "8-K" in form:
        fund = ""
    elif "13F" in form:
        companyName = ""
    elif "13D" in form and ticker != "":
        fund = ""
    elif "13D" in form and ticker == "":
        companyName = ""
    elif "13G" in form and ticker != "":
        fund = ""
    elif "13G" in form and ticker == "":
        companyName = ""
    return filter_nonprintable(ticker), \
           filter_nonprintable(fund), \
           filter_nonprintable(form), \
           filter_nonprintable(purpose), \
           filter_nonprintable(companyName), \
           timeStamp, \
           filter_nonprintable(link), \
           filter_nonprintable(emailBody), \
           filter_nonprintable(longText)


def checkInCsv(dataset, datalist, ticker, fund, form, purpose, companyName, timeStamp, link, emailBody, longText,
               rowFound):
    # check in file
    for ele in datalist:
        if ele[7] == link:
            rowFound = True
            return rowFound
    print(tuple([form, ticker, companyName, fund, purpose, emailBody, longText, link, timeStamp]))
    datalist.insert(0, [form, ticker, companyName, fund, purpose, emailBody, longText, link, timeStamp])
    dataset.add(tuple([form, ticker, companyName, fund, purpose, emailBody, longText, link, timeStamp]))
    sendEmail(form, ticker, companyName, fund, purpose, emailBody, longText, link, timeStamp)


def generate_list(dataset, datalist):
    with open(outputPath + "output.csv", "r") as csvFile:
        reader = csv.reader(csvFile)
        for row in reader:
            dataset.add(tuple(row))
            datalist.append(row)
    return dataset, datalist


# def generateMessage()->MIMEMultipart:
#     message = MIMEMultipart("alternative", None, [MIMEText(HTML, 'html')])
#     message['Subject'] = SUBJECT
#     message['From'] = SENDER_EMAIL
#     message['To'] = RECEIVER_EMAIL
#     return message


def sendEmail(form, ticker, companyName, fund, purpose, emailBody, longText, link, timeStamp):
    import xlrd
    data = []
    cols = []
    wb = xlrd.open_workbook(excelPath + "/" + excelName)
    sheet = wb.sheet_by_index(0)

    for i in range(1, sheet.nrows):
        if ticker == sheet.cell(i, 0).value:
            for j in range(0, 31):
                if str(sheet.cell(0, j).value) == "":
                    cols.append("column" + str(j))
                else:
                    cols.append(str(sheet.cell(0, j).value))
                if str(sheet.cell(i, j).value) == "":
                    data.append("-")
                else:
                    data.append(str(sheet.cell(i, j).value))
            # print(len(data), data)
            # print(len(cols), cols)
            break

    SUBJECT = form + ": " + ticker + " " + companyName + " " + fund + " " + purpose + "\n"

    df = pd.DataFrame(data=[data], columns=cols)
    # print(df)

    if os.path.exists("Table.html"):
        os.remove("Table.html")
    if os.path.exists("Table2.html"):
        os.remove("Table2.html")

    df.to_html("Table.html", index=False)

    df2 = pd.DataFrame(data=[[emailBody], [longText], [link]], columns=["dummy_col"])
    df2.to_html("Table2.html", index=False)
    htmltable2 = df2.to_html(index=False)

    htmltable = df.to_html(index=False)
    HTML = """
    <!DOCTYPE html>
    <html>
      <head>
        <meta charset="utf-8" />
        <style type="text/css">
          table {
            background: white;
            border-radius:3px;
            border-collapse: collapse;
            height: auto;
            max-width: 900px;
            padding:5px;
            width: 100%;
            animation: float 5s infinite;
          }
          th {
            color:#D5DDE5;;
            background:#1b1e24;
            border-bottom: 4px solid #9ea7af;
            font-size:14px;
            font-weight: 300;
            padding:10px;
            text-align:center;
            vertical-align:middle;
          }
          tr {
            border-top: 1px solid #C1C3D1;
            border-bottom: 1px solid #C1C3D1;
            border-left: 1px solid #C1C3D1;
            color:#666B85;
            font-size:16px;
            font-weight:normal;
          }
          tr:hover td {
            background:#4E5066;
            color:#FFFFFF;
            border-top: 1px solid #22262e;
          }
          td {
            background:#FFFFFF;
            padding:10px;
            text-align:left;
            vertical-align:middle;
            font-weight:300;
            font-size:13px;
            border-right: 1px solid #C1C3D1;
          }
        </style>
      </head>
      <body>""" + htmltable2 + htmltable + """</body>
    </html>
    """

    SERVER = str(smtpServerName) + ":" + str(int(smtpServerPort))
    FROM = senderEmailId
    TO = receiver

    message = MIMEMultipart('alternative')
    part2 = MIMEText(HTML, "html")
    # message.attach(part1)
    message.attach(part2)
    # message = MIMEMultipart("alternative", None, [MIMEText(HTML, 'html')])
    message['Subject'] = From + SUBJECT
    message['From'] = FROM
    message['To'] = TO

    # message = """\
    # From: %s
    # To: %s
    # Subject: %s
    #
    # %s
    # """ % (FROM, ", ".join(TO), SUBJECT, TEXT)

    # Send the mail
    server = smtplib.SMTP(SERVER)
    # server.starttls()
    server.login(senderEmailId, str(int(senderEmailPassword)))  # use str(int(pass)) for local server
    server.sendmail(FROM, TO, str(message))
    server.quit()


def getInfo():
    import xlrd
    data = []
    wb = xlrd.open_workbook("Control.xls")
    sheet = wb.sheet_by_index(0)
    for i in range(sheet.nrows):
        data.append(sheet.cell(i, 1).value)
    print(data)
    return data


outputPath, smtpServerName, smtpServerPort, senderEmailId, senderEmailPassword, secure, encryptedConnection, receiver, From, excelPath, excelName = getInfo()
outputPath += "\\"
excelPath += "\\"
